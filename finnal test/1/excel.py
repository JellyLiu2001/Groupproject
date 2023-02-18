import openpyxl
import serial # 导入模块
ydegree=[]
zdegree=[]
angle=[]
wb=openpyxl.Workbook()
ws=wb.create_sheet("sheet1")
ser = serial.Serial('COM8',9600,timeout=1)
# serial.Serial  的三个形参 分别对应 Arduino的串口  波特率 连接超时时间
#print(ser)
while 1:
    val = ser.readline().decode('utf-8')
    # ser.readline() 读取窗串口中的数据以二进制的形式展示需要使用.decode('utf-8')进行解码
    #print(val)
    parsed = val.split(',')
    parsed = [x.strip() for x in parsed]
    #print("parsed",parsed)
    if len(parsed)==3:
        angle.append(int(parsed[0]))
        ydegree.append(float(parsed[1]))
        zdegree.append(float(parsed[2]))
    if parsed==['finish']:
        a=len(angle)
        for i in range(a):
            b=i+1
            ws.cell(row=b, column=1).value=angle[i]
            ws.cell(row=b, column=2).value=ydegree[i]
            ws.cell(row=b, column=3).value=zdegree[i]
        wb.save('data.xlsx')
        print("finnish")
        break;